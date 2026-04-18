"""天堂經典版 LINE 群組聊天記錄分析工具（多廠商）

解析 LINE 外掛交流群的聊天紀錄，自動分類並整理出：
1. 掛機攻略（地點、收益、裝備門檻、職業建議）
2. 設定教學（半自動/全自動設定、各功能設定方法）
3. 帳號安全（封鎖風險、監獄機制、防範建議）
4. 買賣交易（殘卡、序號、天幣、裝備）
5. Bug 排除（閃退、小綠人、連線、卡牆）
6. 環境設定（防毒、遠端、虛擬機、電腦配置）

支援多廠商：透過 --vendor 指定（特工、機器熊等）
"""

import re
import sys
import argparse
from datetime import datetime, timedelta
from pathlib import Path
from collections import defaultdict

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── 重用 analyze.py 的聊天解析器 ───────────────────────────────
from analyze import parse_chat_file
from vendor_config import get_vendor, list_vendors

# ── 系統訊息（需跳過）──────────────────────────────────────────
SKIP_KEYWORDS = ['加入聊天', '已收回訊息', 'Auto-reply', '歡迎您加入']
NOISE_KEYWORDS = ['貼圖', '圖片', '影片']

# ── 分類關鍵字定義 ──────────────────────────────────────────────

# 1. 掛機攻略
FARMING_KEYWORDS = {
    '地點': [
        '古丁', '古1', '古2', '古3', '古4', '古5', '古6',
        '古一', '古二', '古三', '古四', '古五', '古六',
        '蟻洞', '蟻動', '海監', '海因', '眠洞', '棉洞', '棉1', '棉一',
        '墮落', '汙染', '污染', '龍谷', '龍一', '龍六',
        '騎2', '騎3', '騎二', '騎三',
        '話島', '說話之島', '風木', '奇岩',
        '北島', '祝福之地', '骷髏', '野外',
        '網咖', '沙漏',
    ],
    '收益': [
        '一小時', '每小時', '一天', '每天', '24小時',
        r'\d+萬', r'\d+w', 'hr', '收益', '效益',
        '好賺', '賺錢', '打幣', '天幣',
    ],
    '裝備門檻': [
        '防武滿', '防滿', '武滿', '安定', '安定裝',
        r'AC\d+', r'ac\d+', '防禦', '掛得住', '掛不住',
        '會噴水', '喝水', '消耗品',
    ],
    '職業建議': [
        '妖精', '騎士', '法師', '王族',
        '體妖', '敏妖', '體騎', '力騎',
        '好掛', '不好掛', '比較好掛', '比較慘',
    ],
}

# 2. 設定教學
SETTING_KEYWORDS = {
    '半自動設定': [
        '半自動', '手動登',
    ],
    '全自動設定': [
        '全自動', '自動登入', '自動飛', '自動回村',
    ],
    '補血喝水': [
        '補血', '喝水', '喝紅', 'HP', 'hp', '血量',
        '初治', '心靈轉換', '心靈',
    ],
    '順移瞬移': [
        '順移', '瞬移', '無經驗', '無怪', '順飛',
        '飛回', '飛走',
    ],
    '變身設定': [
        '變身', '金變', '普通變身', '金色傳說',
    ],
    '撿物設定': [
        '撿物', '撿東西', '撿錢', '禮儀', '檢物', '減物',
        '簡物', '檢取', '撿取',
    ],
    '組隊BUFF': [
        '組隊', 'buff', 'BUFF', '大地祝福', '大地',
        '舞躍', '大火', '狂暴', '狀態機',
    ],
    '頭盔切換': [
        '換頭盔', '頭盔', '力盔', '加速', '通暢',
    ],
    '洗魔設定': [
        '洗魔', '回魔',
    ],
    '防衛黑名單': [
        '黑名單', '黑單', '反擊', '被打', '被攻擊',
        '白目', '防衛',
    ],
    '巡邏打怪': [
        '巡邏', '巡路', '定點', '攻擊範圍', '自動攻擊',
        '貼怪', '禮貌模式', '戰鬥反應',
    ],
    '回村補給': [
        '回村', '補給', '回城', '存倉', '修武', '修刀',
        '買物', '賣物', '負重',
    ],
}

# 3. 帳號安全
SECURITY_KEYWORDS = {
    '封鎖風險': [
        '被鎖', '被封', '永鎖', '軟鎖', '帳號限制', '鎖帳',
        '畢業', '災情', '制裁',
    ],
    '監獄機制': [
        '監獄', '關監獄', '進監獄', '出獄',
    ],
    '驗證機制': [
        '驗證', '信鴿', '資安', '認證',
    ],
    '排隊問題': [
        '排隊', '排到', '秒進', '沙漏',
    ],
    '防範建議': [
        '鎖一開二', '不要怕', '安心', '風險',
        '低調', '存活率',
    ],
}

# 4. Bug 排除
BUG_KEYWORDS = {
    '閃退問題': [
        '閃退', '跳掉', '跳出', '當掉', '斷線',
        '崩潰', '開不了', '開不起',
    ],
    '小綠人問題': [
        '小綠人', '綠人', '帽子',
    ],
    '連線問題': [
        '連線', '連接', '連不上', '登不上',
        '登入', '無法連',
    ],
    '卡牆發呆': [
        '卡牆', '撞牆', '磨牆', '卡住', '發呆',
        '站在原地', '不會動', '不會打',
    ],
}

# 5. 環境設定
ENV_KEYWORDS = {
    '防毒防火牆': [
        '防毒', '防火牆', 'Defender', 'defender',
    ],
    '遠端監控': [
        '紫P', '紫p', 'Google遠端', 'google遠端', 'goole遠端',
        '遠端', 'OSlink', 'oslink',
    ],
    '虛擬機': [
        '虛擬機', 'VM', 'vm',
    ],
    '電腦配置': [
        'CPU', 'cpu', '記憶體', '顯卡', '99%', '雙開',
    ],
}

# 6. 買賣交易
TRADE_KEYWORDS = {
    '殘卡交易': [
        '殘卡', '收殘', '賣殘', '殘障手冊',
    ],
    '序號買賣': [
        '收卡', '買卡', '賣卡', '新卡', '序號',
        '購買', '購卡',
    ],
    '代設定服務': [
        '代設', '代客', '付費設定', '付費群',
        '腳本設置', '腳本客製',
    ],
    '天幣裝備': [
        r'1[:：]\d+', '天幣', 'MyCard', 'mycard',
        '點數', '儲值',
    ],
}

# 所有分類合併
ALL_CATEGORIES = {
    '掛機攻略': FARMING_KEYWORDS,
    '設定教學': SETTING_KEYWORDS,
    '帳號安全': SECURITY_KEYWORDS,
    'Bug排除': BUG_KEYWORDS,
    '環境設定': ENV_KEYWORDS,
    '買賣交易': TRADE_KEYWORDS,
}

# 廣告暱稱（預設值；會被 vendor_config 覆蓋）
AD_NICKNAMES = [
    '機加酒', '絲瓜蛤蜊', '懶人救星', 'Dong徐',
    '泰雅-腳本設置', '山姆雲科',
]

# 當前廠商設定（由 CLI 設定）
_vendor_cfg = None


def set_vendor(vendor_name):
    """設定當前廠商，載入對應設定。"""
    global _vendor_cfg, AD_NICKNAMES, ALL_CATEGORIES
    _vendor_cfg = get_vendor(vendor_name)
    AD_NICKNAMES = _vendor_cfg.get('ad_nicknames', AD_NICKNAMES)

    # 合併廠商額外關鍵字
    extra = _vendor_cfg.get('extra_keywords', {})
    for main_cat, sub_cats in extra.items():
        if main_cat not in ALL_CATEGORIES:
            ALL_CATEGORIES[main_cat] = {}
        for sub_cat, keywords in sub_cats.items():
            if sub_cat in ALL_CATEGORIES[main_cat]:
                ALL_CATEGORIES[main_cat][sub_cat] = list(
                    dict.fromkeys(ALL_CATEGORIES[main_cat][sub_cat] + keywords)
                )
            else:
                ALL_CATEGORIES[main_cat][sub_cat] = keywords

    return _vendor_cfg


# ── 工具函式 ────────────────────────────────────────────────────

def _is_noise(msg):
    """判斷訊息是否為系統訊息或純噪音。"""
    content = msg.get('content', '')
    raw = msg.get('raw', '')
    text = content if content.strip() else raw

    for kw in SKIP_KEYWORDS:
        if kw in text:
            return True

    # 廠商的 Auto-reply 過濾
    if _vendor_cfg:
        for kw in _vendor_cfg.get('auto_reply_keywords', []):
            if kw in text:
                return True

    # 純貼圖/圖片/影片
    stripped = text.strip()
    if stripped in NOISE_KEYWORDS:
        return True

    # 太短的訊息通常無資訊量
    if len(stripped) < 3:
        return True

    return False


def _is_ad(msg):
    """判斷訊息是否為重複洗版廣告。"""
    nickname = msg.get('nickname', '')
    content = msg.get('content', '')
    for ad in AD_NICKNAMES:
        if ad in nickname:
            return True
    # 代客收費設定的制式廣告
    if '代客收費設定' in content and '串聊' in content:
        return True
    return False


def _match_keywords(text, keywords):
    """檢查文字是否匹配任一關鍵字（支援正則）。回傳匹配到的關鍵字列表。"""
    matched = []
    for kw in keywords:
        if kw.startswith(r'\\') or any(c in kw for c in r'[]+*?{}|()'):
            # 正則模式
            if re.search(kw, text, re.IGNORECASE):
                matched.append(kw)
        else:
            if kw in text:
                matched.append(kw)
    return matched


def classify_message(msg):
    """分類單則訊息。回傳 [(主分類, 子分類, 匹配關鍵字), ...]"""
    content = msg.get('content', '')
    raw = msg.get('raw', '')
    text = content if content.strip() else raw

    results = []
    for main_cat, sub_cats in ALL_CATEGORIES.items():
        for sub_cat, keywords in sub_cats.items():
            matched = _match_keywords(text, keywords)
            if matched:
                results.append((main_cat, sub_cat, matched))

    return results


# ── 分析核心 ────────────────────────────────────────────────────

def analyze_guide(messages, days_limit=None):
    """分析交流群訊息，回傳分類後的結構化資料。

    Returns:
        dict: {
            '掛機攻略': {
                '地點': [{'nickname', 'content', 'date', 'time', 'keywords'}, ...],
                ...
            },
            ...
        }
    """
    cutoff = None
    if days_limit:
        cutoff = datetime.now() - timedelta(days=days_limit)

    categorized = {}
    for main_cat in ALL_CATEGORIES:
        categorized[main_cat] = defaultdict(list)

    for msg in messages:
        if cutoff and msg.get('date') and msg['date'] < cutoff:
            continue

        if _is_noise(msg):
            continue

        if _is_ad(msg):
            continue

        classifications = classify_message(msg)
        if not classifications:
            continue

        for main_cat, sub_cat, matched_kw in classifications:
            categorized[main_cat][sub_cat].append({
                'date': msg['date'].strftime('%Y-%m-%d') if msg.get('date') else '',
                'time': msg.get('time', ''),
                'nickname': msg.get('nickname', ''),
                'content': msg.get('content', ''),
                'keywords': ', '.join(matched_kw),
            })

    # 將 defaultdict 轉為 dict
    return {k: dict(v) for k, v in categorized.items()}


def build_summary(categorized):
    """從分類資料建立摘要統計。"""
    summary = []
    for main_cat, sub_cats in categorized.items():
        total = sum(len(msgs) for msgs in sub_cats.values())
        summary.append({
            '主分類': main_cat,
            '訊息總數': total,
            '子分類數': len(sub_cats),
            '子分類明細': ', '.join(
                f"{k}({len(v)})" for k, v in
                sorted(sub_cats.items(), key=lambda x: -len(x[1]))
            ),
        })
    summary.sort(key=lambda x: -x['訊息總數'])
    return pd.DataFrame(summary)


def _deduplicate_tips(messages):
    """對同義/重複的訊息進行去重，保留有價值的。"""
    seen_content = set()
    unique = []
    for msg in messages:
        # 用內容的前 30 字去重
        key = msg['content'][:30].strip()
        if key and key not in seen_content:
            seen_content.add(key)
            unique.append(msg)
    return unique


def extract_farming_tips(categorized):
    """提取掛機攻略的精華內容。"""
    farming = categorized.get('掛機攻略', {})
    tips = []

    for sub_cat, messages in farming.items():
        deduped = _deduplicate_tips(messages)
        for msg in deduped:
            content = msg['content']
            # 過濾純問句（以？結尾且很短）
            if content.strip().endswith('?') or content.strip().endswith('？'):
                if len(content) < 20:
                    continue
            tips.append({
                '子分類': sub_cat,
                '日期': msg['date'],
                '時間': msg['time'],
                '發言者': msg['nickname'],
                '內容': content,
                '關鍵字': msg['keywords'],
            })

    return pd.DataFrame(tips)


def extract_setting_tips(categorized):
    """提取設定教學的精華內容。"""
    settings = categorized.get('設定教學', {})
    tips = []

    for sub_cat, messages in settings.items():
        deduped = _deduplicate_tips(messages)
        for msg in deduped:
            content = msg['content']
            if content.strip().endswith('?') or content.strip().endswith('？'):
                if len(content) < 20:
                    continue
            tips.append({
                '子分類': sub_cat,
                '日期': msg['date'],
                '時間': msg['time'],
                '發言者': msg['nickname'],
                '內容': content,
                '關鍵字': msg['keywords'],
            })

    return pd.DataFrame(tips)


# ── Excel 輸出 ──────────────────────────────────────────────────

def _style_sheet(ws, header_fill_color='4472C4'):
    """統一設定工作表樣式。"""
    header_fill = PatternFill(start_color=header_fill_color, end_color=header_fill_color, fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )

    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border
            if cell.row > 1:
                cell.alignment = Alignment(vertical='top', wrap_text=True)

    # 自動調整欄寬
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 50), min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 60)


def export_excel(categorized, output_path):
    """將分析結果匯出到 Excel。"""
    # 分類顏色
    cat_colors = {
        '掛機攻略': '2E7D32',  # 深綠
        '設定教學': '1565C0',  # 深藍
        '帳號安全': 'C62828',  # 深紅
        'Bug排除': 'E65100',  # 深橙
        '環境設定': '6A1B9A',  # 深紫
        '買賣交易': '4E342E',  # 深棕
    }

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # 1. 摘要總覽
        summary_df = build_summary(categorized)
        if not summary_df.empty:
            summary_df.to_excel(writer, sheet_name='摘要總覽', index=False)
            _style_sheet(writer.sheets['摘要總覽'], '333333')

        # 2. 掛機攻略（優先）
        farming_df = extract_farming_tips(categorized)
        if not farming_df.empty:
            farming_df.to_excel(writer, sheet_name='掛機攻略', index=False)
            _style_sheet(writer.sheets['掛機攻略'], cat_colors['掛機攻略'])

        # 3. 設定教學
        setting_df = extract_setting_tips(categorized)
        if not setting_df.empty:
            setting_df.to_excel(writer, sheet_name='設定教學', index=False)
            _style_sheet(writer.sheets['設定教學'], cat_colors['設定教學'])

        # 4. 其他分類各自一個 sheet
        for main_cat in ['帳號安全', 'Bug排除', '環境設定', '買賣交易']:
            cat_data = categorized.get(main_cat, {})
            rows = []
            for sub_cat, messages in cat_data.items():
                for msg in _deduplicate_tips(messages):
                    rows.append({
                        '子分類': sub_cat,
                        '日期': msg['date'],
                        '時間': msg['time'],
                        '發言者': msg['nickname'],
                        '內容': msg['content'],
                        '關鍵字': msg['keywords'],
                    })
            if rows:
                df = pd.DataFrame(rows)
                df.to_excel(writer, sheet_name=main_cat, index=False)
                _style_sheet(
                    writer.sheets[main_cat],
                    cat_colors.get(main_cat, '4472C4'),
                )

    return output_path


# ── CLI ─────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description='天堂經典版 LINE 外掛交流群聊天記錄分析工具（多廠商）',
    )
    parser.add_argument(
        '--vendor', '-v',
        type=str,
        default=None,
        help=f'指定廠商名稱（可用：{", ".join(list_vendors())}）',
    )
    parser.add_argument(
        'input_file',
        type=str,
        nargs='?',
        default=None,
        help='LINE 聊天記錄 .txt 檔案路徑（若指定 --vendor 可省略）',
    )
    parser.add_argument(
        '-o', '--output',
        type=str,
        default=None,
        help='輸出 Excel 檔案路徑（預設：out/<廠商>/<廠商>_分析.xlsx）',
    )
    parser.add_argument(
        '-d', '--days',
        type=int,
        default=None,
        help='僅分析最近 N 天的訊息（預設：全部）',
    )
    parser.add_argument(
        '--summary',
        action='store_true',
        help='僅在終端顯示摘要（不匯出 Excel）',
    )
    parser.add_argument(
        '--list-vendors',
        action='store_true',
        help='列出所有可用廠商',
    )

    args = parser.parse_args()

    # ── 列出廠商 ──
    if args.list_vendors:
        print('可用廠商：')
        for v in list_vendors():
            print(f'  {v}')
        return

    # ── 載入廠商設定 ──
    vendor_name = None
    if args.vendor:
        vendor_cfg = set_vendor(args.vendor)
        vendor_name = vendor_cfg['name']
        print(f'廠商：{vendor_name}（{vendor_cfg["full_name"]}）')

    # ── 決定輸入檔案 ──
    if args.input_file:
        input_path = Path(args.input_file)
        if not input_path.exists():
            print(f'錯誤：找不到檔案 {input_path}')
            sys.exit(1)
        files = [input_path]
    elif vendor_name and _vendor_cfg:
        txt_dir = _vendor_cfg['txt_dir']
        if not txt_dir.exists():
            print(f'錯誤：找不到廠商聊天資料夾 {txt_dir}')
            sys.exit(1)
        files = sorted(txt_dir.glob('*.txt'))
        if not files:
            print(f'錯誤：{txt_dir} 中沒有 .txt 檔案')
            sys.exit(1)
    else:
        print('錯誤：請指定 --vendor 或提供 input_file')
        sys.exit(1)

    # ── 解析 ──
    all_msgs = []
    for fp in files:
        print(f'解析聊天記錄：{fp.name} ...')
        msgs = parse_chat_file(str(fp))
        print(f'  共 {len(msgs)} 則訊息')
        all_msgs.extend(msgs)

    if len(files) > 1:
        print(f'合計 {len(all_msgs)} 則訊息')

    # ── 分析 ──
    print('分類分析中 ...')
    categorized = analyze_guide(all_msgs, days_limit=args.days)

    # ── 摘要 ──
    summary_df = build_summary(categorized)
    label = f'  {vendor_name} 分類摘要' if vendor_name else '  分類摘要'
    print('\n' + '=' * 60)
    print(label)
    print('=' * 60)
    for _, row in summary_df.iterrows():
        print(f"  {row['主分類']:<8}  {row['訊息總數']:>4} 則  {row['子分類明細']}")
    print('=' * 60)

    if args.summary:
        return

    # ── 匯出 ──
    if args.output:
        out_path = Path(args.output)
    elif vendor_name and _vendor_cfg:
        out_dir = _vendor_cfg['out_dir']
        out_dir.mkdir(parents=True, exist_ok=True)
        out_path = out_dir / f'{vendor_name}_分析.xlsx'
    else:
        out_path = files[0].with_name(files[0].stem + '_guide_analysis.xlsx')

    out_path.parent.mkdir(parents=True, exist_ok=True)
    export_excel(categorized, str(out_path))
    print(f'\n已匯出：{out_path}')


if __name__ == '__main__':
    main()
