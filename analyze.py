"""天堂經典版 LINE 群組交易分析工具

解析 LINE 聊天記錄匯出的 .txt 檔，擷取買賣訊息並匯出 Excel 報表。
"""

import re
import sys
import shutil
import argparse
from datetime import datetime, timedelta
from pathlib import Path

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── 物品別名對照表 ──────────────────────────────────────────────
ITEM_ALIASES = {
    '大馬士革': ['大馬士革', '大馬'],
    '艾爾穆': ['艾爾穆', '艾盔'],
    '精靈內衣': ['精靈內衣', '精內'],
    '保斗': ['保斗', '寶斗', '保抖'],
    '鋼手': ['鋼手', '剛手'],
    '鋼靴': ['鋼靴', '剛靴', '剛薛'],
    '精盾': ['精盾'],
    '精甲': ['精甲'],
    '紅頭巾': ['紅頭巾', '紅頭'],
    '鎖破': ['鎖破'],
    '十字': ['十字'],
    '尤米': ['尤米'],
    '武士刀': ['武士刀'],
    '力杖': ['力杖', '力仗'],
    '精煉': ['精煉', '精鏈', '精T'],
    '腕甲': ['腕甲'],
    '抗盔': ['抗盔', '抗頭', '抗虧'],
    '多羅皮帶': ['多羅皮帶', '多羅'],
    '武卷': ['武卷'],
    '防卷': ['防卷'],
    '品鑽': ['品鑽', '品質鑽石'],
    '天幣': ['天幣'],
    '序號': ['序號'],
    '反盾': ['反盾'],
    '細劍': ['細劍'],
    '棉袍': ['棉袍', '棉質長袍'],
    '絲質長袍': ['絲質長袍'],
    '夏納變身卷': ['夏納變身', '夏納變捲', '夏納'],
    '冥想': ['冥想'],
    '龍捲風': ['龍捲風'],
    '雙手劍': ['雙手劍'],
    '祝防': ['祝防', '祝福防具'],
    '祝武': ['祝武', '祝福武器'],
    '守護戒指': ['守護戒指'],
    '妖魔項鍊': ['妖魔項鍊'],
    '智力項鍊': ['智力項鍊', '智鍊'],
    '敏捷頭盔': ['敏捷頭盔', '敏盔'],
    '體魄法書': ['體魄法書', '體魄'],
}

# 建立反向對照：別名 → 標準名稱（長的優先以利貪婪比對）
ALIAS_TO_ITEM = {}
for _std, _aliases in ITEM_ALIASES.items():
    for _a in _aliases:
        ALIAS_TO_ITEM[_a] = _std
SORTED_ALIASES = sorted(ALIAS_TO_ITEM.keys(), key=len, reverse=True)

# 系統訊息關鍵字：出現這些表示非交易訊息
SKIP_KEYWORDS = ['加入聊天', '已收回訊息', 'Auto-reply', '歡迎您加入', '貼圖', '圖片']


# ── 解析 ──────────────────────────────────────────────────────────
def parse_chat_file(file_path):
    """將 LINE 聊天匯出 .txt 解析為結構化訊息列表。"""
    with open(file_path, 'r', encoding='utf-8-sig') as f:
        lines = f.readlines()

    messages = []
    current_date = None
    current_msg = None

    for line in lines:
        line = line.rstrip('\n').rstrip('\r')

        # 日期行：2026.02.26 星期四
        date_match = re.match(r'^(\d{4}\.\d{2}\.\d{2})\s', line)
        if date_match:
            if current_msg:
                messages.append(current_msg)
                current_msg = None
            current_date = datetime.strptime(date_match.group(1), '%Y.%m.%d')
            continue

        # 訊息行：HH:MM 暱稱 內容
        msg_match = re.match(r'^(\d{2}:\d{2})\s+(.+)', line)
        if msg_match:
            if current_msg:
                messages.append(current_msg)
            rest = msg_match.group(2)
            # 嘗試分離暱稱與訊息（以第一個空白切割）
            parts = rest.split(None, 1)
            nickname = parts[0] if parts else ''
            msg_body = parts[1] if len(parts) > 1 else ''
            current_msg = {
                'date': current_date,
                'time': msg_match.group(1),
                'nickname': nickname,
                'content': msg_body,  # 僅訊息本體（不含暱稱）
                'raw': rest,          # 含暱稱的完整行
            }
        else:
            # 多行訊息的後續行
            if current_msg and line.strip():
                current_msg['content'] += ' ' + line.strip()

    if current_msg:
        messages.append(current_msg)

    return messages


def detect_action(content):
    """辨識買或賣動作。回傳 '收購' / '出售' / '買賣皆有' / None。"""
    for kw in SKIP_KEYWORDS:
        if kw in content:
            return None

    # 排除常見誤判：收到、收回、買到了 等日常用語
    cleaned = re.sub(r'收到|收回|收入|買到|買不到|出來|出去|出現|出了|出發|出生', '', content)

    has_buy = bool(re.search(r'收(?![到回入藏集])|買', cleaned))
    has_sell = bool(re.search(r'賣|售|出(?:天幣|幣|金|\d)', cleaned))

    if has_buy and has_sell:
        return '買賣皆有'
    if has_buy:
        return '收購'
    if has_sell:
        return '出售'
    return None


def detect_items(content):
    """偵測訊息中的物品。回傳 [(標準名稱, 強化等級), ...]。"""
    found = []
    for alias in SORTED_ALIASES:
        if alias not in content:
            continue
        std = ALIAS_TO_ITEM[alias]
        if any(s == std for s, _ in found):
            continue
        # 嘗試抓 +N 強化等級
        pat = re.search(r'\+(\d{1,2})\s*' + re.escape(alias), content)
        if pat:
            enh = f"+{pat.group(1)}"
        else:
            # 也抓「7艾盔」這種省略 + 號的寫法
            pat2 = re.search(r'(?<!\d)(\d{1,2})\s*' + re.escape(alias), content)
            if pat2 and int(pat2.group(1)) <= 10:
                enh = f"+{pat2.group(1)}"
            else:
                enh = ''
        found.append((std, enh))
    return found


def detect_prices(content):
    """擷取價格。回傳 [(值, 單位), ...]。"""
    prices = []
    # 匯率 1:175
    for m in re.findall(r'1\s*[:：]\s*(\d{2,3})', content):
        prices.append((f'1:{m}', '匯率'))
    # 數字T
    for m in re.findall(r'(\d+\.?\d*)\s*[Tt](?![a-zA-Z])', content):
        prices.append((m, 'T'))
    # 數字萬 或 數字W
    for m in re.findall(r'(\d+\.?\d*)\s*(?:萬|[wW])', content):
        prices.append((m, '萬'))
    return prices


# ── 分析 ──────────────────────────────────────────────────────────
def analyze(messages, days_limit, target_items=None, skip_other=True):
    """分析訊息列表，回傳交易 DataFrame。"""
    cutoff = datetime.now() - timedelta(days=days_limit)
    rows = []

    for msg in messages:
        if not msg['date'] or msg['date'] < cutoff:
            continue

        content = msg.get('content', '')
        raw = msg.get('raw', content)
        nickname = msg.get('nickname', '')

        # 優先分析訊息本體，若本體空白則回退到完整行
        text = content if content.strip() else raw
        action = detect_action(text)
        if not action:
            continue

        # 物品和價格也從訊息本體+暱稱中抓（暱稱本身可能含交易資訊）
        items = detect_items(text)
        prices = detect_prices(text)

        # 暱稱中的物品/價格也補充進來（如「售+5剛靴4700t」這種暱稱）
        nick_items = detect_items(nickname)
        nick_prices = detect_prices(nickname)
        for ni in nick_items:
            if not any(ni[0] == i[0] for i in items):
                items.append(ni)
        for np_ in nick_prices:
            if np_ not in prices:
                prices.append(np_)

        if not items:
            if skip_other:
                continue
            items = [('其他', '')]

        if target_items:
            items = [(n, e) for n, e in items if n in target_items]
            if not items:
                continue

        price_str = ' / '.join(f'{v} {u}' for v, u in prices) if prices else '面議/私訊'

        for item_name, enh in items:
            display = f'{enh} {item_name}'.strip() if enh else item_name
            rows.append({
                '日期': msg['date'].strftime('%Y-%m-%d'),
                '時間': msg['time'],
                '物品': display,
                '物品分類': item_name,
                '強化等級': enh,
                '動作': action,
                '價格': price_str,
                '完整訊息': raw,
            })

    return pd.DataFrame(rows)


def _extract_numeric_prices(price_series):
    """從價格字串中提取數值，回傳 {單位: [數值列表]}。"""
    result = {'T': [], '萬': []}
    for ps in price_series:
        if not isinstance(ps, str):
            continue
        for m in re.findall(r'(\d+\.?\d*)\s*T', ps):
            try:
                result['T'].append(float(m))
            except ValueError:
                pass
        for m in re.findall(r'(\d+\.?\d*)\s*萬', ps):
            try:
                result['萬'].append(float(m))
            except ValueError:
                pass
    return result


def _price_range_str(values):
    """格式化價格範圍。"""
    if not values:
        return '-'
    values = sorted(values)
    if len(values) == 1:
        return f'{values[0]:g}'
    return f'{values[0]:g} ~ {values[-1]:g}'


def build_summary(df):
    """建立行情總覽摘要表。"""
    if df.empty:
        return pd.DataFrame()

    rows = []
    for item in df['物品'].unique():
        sub = df[df['物品'] == item]
        buy = sub[sub['動作'].isin(['收購', '買賣皆有'])]
        sell = sub[sub['動作'].isin(['出售', '買賣皆有'])]

        buy_prices = _extract_numeric_prices(buy['價格'])
        sell_prices = _extract_numeric_prices(sell['價格'])

        # 去重筆數（相同完整訊息視為重複張貼）
        unique_buy = buy['完整訊息'].nunique() if len(buy) else 0
        unique_sell = sell['完整訊息'].nunique() if len(sell) else 0

        rows.append({
            '物品': item,
            '收購筆數': len(buy),
            '出售筆數': len(sell),
            '去重收購': unique_buy,
            '去重出售': unique_sell,
            '收購價(T)': _price_range_str(buy_prices['T']),
            '收購價(萬)': _price_range_str(buy_prices['萬']),
            '出售價(T)': _price_range_str(sell_prices['T']),
            '出售價(萬)': _price_range_str(sell_prices['萬']),
            '日期範圍': f"{sub['日期'].min()} ~ {sub['日期'].max()}",
            '總筆數': len(sub),
        })

    return pd.DataFrame(rows).sort_values('總筆數', ascending=False).reset_index(drop=True)


def build_exchange_rate_sheet(df):
    """從天幣相關交易中提取匯率趨勢。"""
    rate_rows = []
    for _, row in df.iterrows():
        if '匯率' not in row['價格']:
            continue
        for m in re.findall(r'1:(\d{2,3})', row['價格']):
            rate_rows.append({
                '日期': row['日期'],
                '時間': row['時間'],
                '動作': row['動作'],
                '匯率': int(m),
                '完整訊息': row['完整訊息'],
            })
    return pd.DataFrame(rate_rows)


# ── Excel 輸出 ────────────────────────────────────────────────────
def _style_sheet(worksheet, df):
    """套用 Excel 樣式。"""
    hdr_fill = PatternFill('solid', fgColor='4472C4')
    hdr_font = Font(name='Microsoft JhengHei', bold=True, color='FFFFFF', size=11)
    buy_fill = PatternFill('solid', fgColor='E2EFDA')
    sell_fill = PatternFill('solid', fgColor='FCE4EC')
    both_fill = PatternFill('solid', fgColor='FFF3E0')
    data_font = Font(name='Microsoft JhengHei', size=10)
    thin = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )

    # 表頭
    for c, col_name in enumerate(df.columns, 1):
        cell = worksheet.cell(1, c)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin

    # 動作欄位索引
    act_idx = list(df.columns).index('動作') + 1 if '動作' in df.columns else None

    # 資料列
    for r in range(2, len(df) + 2):
        fill = None
        if act_idx:
            v = worksheet.cell(r, act_idx).value
            if v == '收購':
                fill = buy_fill
            elif v == '出售':
                fill = sell_fill
            elif v == '買賣皆有':
                fill = both_fill
        for c in range(1, len(df.columns) + 1):
            cell = worksheet.cell(r, c)
            cell.font = data_font
            cell.border = thin
            cell.alignment = Alignment(vertical='center', wrap_text=True)
            if fill:
                cell.fill = fill

    # 欄寬自動調整
    for c, col_name in enumerate(df.columns, 1):
        max_w = len(str(col_name)) * 2.2
        for r in range(2, min(len(df) + 2, 200)):
            val = str(worksheet.cell(r, c).value or '')
            max_w = max(max_w, min(len(val) * 1.3, 55))
        worksheet.column_dimensions[get_column_letter(c)].width = max(max_w, 8)

    worksheet.freeze_panes = 'A2'
    worksheet.auto_filter.ref = worksheet.dimensions


def _sanitize_text(val):
    """移除 XML 不允許的控制字元。"""
    if not isinstance(val, str):
        return val
    return re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f]', '', val)


def export_excel(df, summary_df, rate_df, output_path):
    """匯出多分頁 Excel。"""
    # 清理文字中的非法 XML 字元
    str_cols = df.select_dtypes(include=['object', 'string']).columns
    df = df.copy()
    for col in str_cols:
        df[col] = df[col].map(_sanitize_text)
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        if not summary_df.empty:
            summary_df.to_excel(writer, sheet_name='行情總覽', index=False)
            _style_sheet(writer.sheets['行情總覽'], summary_df)
        if not df.empty:
            df.to_excel(writer, sheet_name='交易明細', index=False)
            _style_sheet(writer.sheets['交易明細'], df)
        if not rate_df.empty:
            rate_df.to_excel(writer, sheet_name='天幣匯率', index=False)
            _style_sheet(writer.sheets['天幣匯率'], rate_df)


# ── 主程式 ────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(
        description='天堂經典版 LINE 群組交易分析工具',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=textwrap(),
    )
    parser.add_argument('--days', type=int, default=30,
                        help='篩選最近 N 天的資料 (預設: 30)')
    parser.add_argument('--input', type=str, default=None,
                        help='指定輸入的 .txt 檔案路徑 (預設: 讀取 txt/ 全部)')
    parser.add_argument('--items', nargs='+', default=None,
                        help='只分析指定物品，如: --items 大馬士革 鋼手')
    parser.add_argument('--rate', type=float, default=None,
                        help='天幣匯率，如: --rate 170 (1T=170萬天幣)')
    parser.add_argument('--output', type=str, default=None,
                        help='指定輸出檔名 (預設: out/天堂交易分析_最新.xlsx)')
    parser.add_argument('--keep', type=int, default=5,
                        help='保留最近幾份存檔 (預設: 5)')
    parser.add_argument('--all', action='store_true',
                        help='包含無法辨識物品的「其他」類別 (預設排除)')
    parser.add_argument('--list-items', action='store_true',
                        help='列出所有可搜尋的物品名稱')

    args = parser.parse_args()
    base = Path(__file__).resolve().parent
    txt_dir = base / 'txt'
    out_dir = base / 'out'

    # ── 列出可搜尋物品 ──
    if args.list_items:
        print('可搜尋物品清單（--items 可用的名稱）：')
        for std, aliases in sorted(ITEM_ALIASES.items()):
            aka = ', '.join(a for a in aliases if a != std)
            extra = f'  (別名: {aka})' if aka else ''
            print(f'  {std}{extra}')
        return

    out_dir.mkdir(exist_ok=True)

    # ── 收集輸入檔案 ──
    if args.input:
        p = Path(args.input)
        if not p.is_absolute():
            p = base / p
        if not p.exists():
            print(f'[錯誤] 找不到檔案: {p}')
            sys.exit(1)
        files = [p]
    else:
        if not txt_dir.exists():
            txt_dir.mkdir()
            print(f'[提示] 已建立 txt/ 資料夾，請將 LINE 聊天匯出 .txt 放入後重新執行。')
            sys.exit(0)
        files = sorted(txt_dir.glob('*.txt'))
        # 也搜尋子資料夾（多廠商結構）
        if not files:
            files = sorted(txt_dir.rglob('*.txt'))
        if not files:
            print('[錯誤] txt/ 資料夾中沒有 .txt 檔案')
            sys.exit(1)

    # ── 解析 ──
    print(f'讀取 {len(files)} 個檔案 ...')
    all_msgs = []
    for fp in files:
        print(f'  -> {fp.name}')
        all_msgs.extend(parse_chat_file(fp))
    print(f'共解析 {len(all_msgs)} 筆訊息')

    # ── 對照 target items ──
    target = None
    if args.items:
        target = set()
        for it in args.items:
            if it in ALIAS_TO_ITEM:
                target.add(ALIAS_TO_ITEM[it])
            elif it in ITEM_ALIASES:
                target.add(it)
            else:
                print(f'  [警告] 未知物品「{it}」，將照原名搜尋')
                target.add(it)

    # ── 分析 ──
    print(f'篩選最近 {args.days} 天 ...')
    skip_other = not args.all
    df = analyze(all_msgs, args.days, target, skip_other=skip_other)

    if df.empty:
        print('[結果] 在指定條件下沒有找到交易訊息。')
        print('  建議：增加 --days 天數、放入更新的 txt、或調整 --items。')
        sys.exit(0)

    # ── 匯率換算欄 ──
    if args.rate:
        rate = args.rate

        def convert(row):
            prices = detect_prices(row['完整訊息'])
            parts = []
            for val, unit in prices:
                try:
                    v = float(val.replace('1:', ''))
                except ValueError:
                    continue
                if unit == 'T':
                    parts.append(f'{v * rate:.1f} 萬天幣')
                elif unit == '萬':
                    parts.append(f'{v / rate:.2f} T')
            return ' / '.join(parts) if parts else ''

        df['匯率換算參考'] = df.apply(convert, axis=1)
        print(f'匯率換算: 1T = {rate} 萬天幣')

    summary_df = build_summary(df)
    rate_df = build_exchange_rate_sheet(df)

    # ── 輸出路徑 ──
    if args.output:
        out_path = Path(args.output)
        if not out_path.is_absolute():
            out_path = out_dir / out_path
    else:
        out_path = out_dir / '天堂交易分析_最新.xlsx'

    # 時間戳記存檔
    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    archive_path = out_dir / f'天堂交易分析_{ts}.xlsx'

    # 清理舊存檔，只保留最近 N 份
    archives = sorted(out_dir.glob('天堂交易分析_2*.xlsx'))
    while len(archives) >= args.keep:
        old = archives.pop(0)
        old.unlink()
        print(f'  [清理] 刪除舊存檔: {old.name}')

    # 寫出 (只產生一次，再複製為存檔)
    export_excel(df, summary_df, rate_df, out_path)
    shutil.copy2(out_path, archive_path)

    # ── 結果摘要 ──
    print()
    print('=' * 55)
    print('  分析完成')
    print('=' * 55)
    print(f'  資料範圍 : 最近 {args.days} 天')
    print(f'  交易筆數 : {len(df)} 筆')
    print(f'  物品種類 : {df["物品分類"].nunique()} 種')
    if not rate_df.empty:
        print(f'  匯率資料 : {len(rate_df)} 筆')
    print()
    print('  物品交易統計 (前 15 名):')
    print(f'  {"物品":<14s} {"收購":>5s} {"出售":>5s} {"收購價(T)":>14s} {"出售價(T)":>14s}')
    print(f'  {"-"*14} {"-"*5} {"-"*5} {"-"*14} {"-"*14}')
    for _, r in summary_df.head(15).iterrows():
        bt = r.get('收購價(T)', '-')
        st = r.get('出售價(T)', '-')
        print(f'  {r["物品"]:<14s} {r["收購筆數"]:5d} {r["出售筆數"]:5d} {str(bt):>14s} {str(st):>14s}')
    print()
    print(f'  最新檔案 : {out_path}')
    print(f'  存檔副本 : {archive_path}')
    print('=' * 55)


def textwrap():
    return """
使用範例:
  python analyze.py                           預設: txt/ 全部, 最近 30 天
  python analyze.py --days 60                 最近 60 天
  python analyze.py --days 90 --items 大馬 鋼手   篩選指定物品
  python analyze.py --input txt/某檔案.txt      指定單一檔案
  python analyze.py --rate 170                匯率換算 (1T = 170萬天幣)
  python analyze.py --list-items              列出所有可搜尋物品
  python analyze.py --keep 3                  只保留最近 3 份存檔
"""


if __name__ == '__main__':
    main()
